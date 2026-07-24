import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
from datetime import datetime

NAME_MAP = {
    "Amhara Bank":    "Amhara",
    "AmharaEthBirr":  "AmharaEthBirr",
    "Awash Bank":     "Awash",
    "Dashen Bank":    "Dashen",
    "COOP":           "CBO",
    "Zemen Bank":     "Zemen",
    "CBE":            "CBE",
    "KAAFI":          "KAAFI MFI",
    "Tsehay Bank":    "Tsehay",
    "Nib Bank":       "NIB",
    "Addis bank":     "Addis",
    "Wegagen Bank":   "Wegagen",
    "Abay Bank":      "Abay",
    "BOA":            "Abyssinia",
    "Buna bank":      "Buna Bank",
    "Rays":           "Rays",
    "Siinqee Bank":   "Siinqee",
    "Wegagen E-Birr": "Wegagen e-Birr",
    "CBEBirr":        "CBE BIRR",
    "Coopay-E-Birr":  "Coopay-E-Birr",
    "MPESA":          "MPESA",
    "Kacha":          "Kacha",
    "Tsedey Bank":    "Tsedey Bank",
    "LIB":            "LIB",
    "ZamZam":         "ZamZam",
    "Enat Bank":      "Enat Bank",
    "Siket":          "Siket",
    "Hijra Bank":     "Hijra Bank",
    "GOH BETOCH":     "GOH BETOCH",
    "Rammis Bank":    "Rammis Bank",
    "Saha":           "Saha",
    "HalalPay":       "HalalPay",
    "Hibret":         "Hibret",
    "Global":         "Global",
    "Vision Fund":    "Vision Fund",
    "Berhan Bank":    "Berhan Bank",
    "Oromia bank":    "Oromia bank",
    "Gadaa":          "Gadaa",
    "Sidama Bank":    "Sidama Bank",
    "Ahadu":          "Ahadu",
    "Ahadu Ebirr":    "Ahadu Ebirr",
    "Nisir":          "Nisir",
    "Shabelle":       "Shabelle",
    "Omo":            "Omo Bank",
    "NibBirr":        "NibE Birr",
    "Dedebit":        "Dedebit MFI",
    "H-Cash":         "H-Cash",
    "Vitabirr":       "Vitabirr",
    "Dire MFI":       "Dire MFI",
    "telebirr":       "telebirr",
}

PERMANENT_FIS = [
    "AmharaEthBirr", "Coopay-E-Birr", "GOH BETOCH", "Hibret",
    "Kacha", "LIB", "Rammis Bank", "Saha", "HalalPay",
    "Tsedey Bank", "ZamZam", "Hijra Bank", "Rays", "Shabelle",
]

GREEN = "A9D08E"
WHITE = "FFFFFF"
BLACK = "000000"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def generate_report(input_path, output_path, report_date):
    df = pd.read_excel(input_path)

    unmapped = []
    rows = []
    seen_names = set()

    for _, r in df.iterrows():
        raw_name = str(r["BANK_NAME"]).strip()
        display_name = NAME_MAP.get(raw_name)
        if display_name is None:
            unmapped.append(raw_name)
            display_name = raw_name

        iss_cnt = int(r["ISSUER_TXN_COUNT"])       if pd.notna(r["ISSUER_TXN_COUNT"])     else 0
        iss_amt = float(r["ISSUER_TOTAL_AMOUNT"])   if pd.notna(r["ISSUER_TOTAL_AMOUNT"])  else 0.0
        acq_cnt = int(r["ACQUIRER_TXN_COUNT"])      if pd.notna(r["ACQUIRER_TXN_COUNT"])   else 0
        acq_amt = float(r["ACQUIRER_TOTAL_AMOUNT"]) if pd.notna(r["ACQUIRER_TOTAL_AMOUNT"]) else 0.0

        rows.append({
            "name": display_name,
            "iss_cnt": iss_cnt, "iss_amt": iss_amt,
            "acq_cnt": acq_cnt, "acq_amt": acq_amt,
        })
        seen_names.add(display_name)

        # Auto-promote to permanent list
        if display_name not in PERMANENT_FIS:
            PERMANENT_FIS.append(display_name)

    # Add permanent FIs with no data today → 0,0,0,0
    for fi in PERMANENT_FIS:
        if fi not in seen_names:
            rows.append({"name": fi, "iss_cnt": 0, "iss_amt": 0.0, "acq_cnt": 0, "acq_amt": 0.0})

    rows.sort(key=lambda x: x["name"].strip().lower())

    if unmapped:
        print(f"\n⚠️  NEW FI(s) DETECTED: {unmapped} — added to report automatically.\n")
    else:
        print("✅ All FIs matched.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    date_str = report_date.strftime("%B %d,%Y")

    ws.merge_cells("B1:F1"); ws["B1"] = "EthSwitch S.C."
    ws["B1"].font = Font(bold=True, size=12); ws["B1"].alignment = Alignment(horizontal="center")
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}1"].border = thin_border()

    ws.merge_cells("B2:F2"); ws["B2"] = "IPS Successful Report"
    ws["B2"].font = Font(bold=True, size=12); ws["B2"].alignment = Alignment(horizontal="center")
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}2"].border = thin_border()

    ws.merge_cells("B3:F3"); ws["B3"] = report_date
    ws["B3"].number_format = "d-mmm-yyyy"
    ws["B3"].font = Font(bold=True, size=12); ws["B3"].alignment = Alignment(horizontal="center")
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}3"].border = thin_border()

    ws.merge_cells("B4:F4")
    ws["B4"] = f"Successful IPS Interoperable Transactions Held on {date_str}"
    ws["B4"].font = Font(bold=True, size=11, color=BLACK)
    ws["B4"].fill = make_fill(WHITE); ws["B4"].alignment = Alignment(horizontal="center")
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}4"].border = thin_border()

    ws.merge_cells("B5:B6"); ws["B5"] = "BANK"
    ws.merge_cells("C5:D5"); ws["C5"] = "Successful  Transactions As Destination"
    ws.merge_cells("E5:F5"); ws["E5"] = "Successful Transactions As Source"
    ws["C6"] = "No.Transactions"; ws["D6"] = "Values"
    ws["E6"] = "No.Transactions"; ws["F6"] = "Values"

    for row in [5, 6]:
        for col in ["B", "C", "D", "E", "F"]:
            cell = ws[f"{col}{row}"]
            cell.fill = make_fill(GREEN)
            cell.font = Font(size=11, color=BLACK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

    data_start = 7
    for i, row in enumerate(rows):
        r = data_start + i
        ws[f"B{r}"] = row["name"]
        ws[f"B{r}"].fill = make_fill(GREEN)
        ws[f"B{r}"].font = Font(size=11, color=BLACK)
        ws[f"B{r}"].border = thin_border()

        for col, val, fmt in [
            ("C", row["iss_cnt"], '#,##0'),
            ("D", row["iss_amt"], '#,##0.00'),
            ("E", row["acq_cnt"], '#,##0'),
            ("F", row["acq_amt"], '#,##0.00'),
        ]:
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = Font(size=11)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = fmt

    total_row = data_start + len(rows)
    last_data = total_row - 1
    ws[f"B{total_row}"] = "TOTAL"
    ws[f"B{total_row}"].fill = make_fill(GREEN)
    ws[f"B{total_row}"].font = Font(bold=True, size=11, color=BLACK, underline="single")
    ws[f"B{total_row}"].border = thin_border()

    for col, fmt in [("C", '#,##0'), ("D", '#,##0.00'), ("E", '#,##0'), ("F", '#,##0.00')]:
        cell = ws[f"{col}{total_row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{last_data})"
        cell.font = Font(bold=True, size=11, underline="single")
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = fmt

    ws.column_dimensions["A"].width = 9.0
    ws.column_dimensions["B"].width = 19.43
    ws.column_dimensions["C"].width = 21.14
    ws.column_dimensions["D"].width = 26.45
    ws.column_dimensions["E"].width = 19.0
    ws.column_dimensions["F"].width = 32.27
    ws.row_dimensions[1].height = 15.6
    ws.row_dimensions[2].height = 15.6
    ws.row_dimensions[3].height = 15.6
    ws.row_dimensions[4].height = 20.25
    ws.row_dimensions[5].height = 13.9

    wb.save(output_path)
    print(f"✅ IPS Report saved → {output_path}")

if __name__ == "__main__":
    input_file  = sys.argv[1]
    date_arg    = sys.argv[2] if len(sys.argv) > 2 else datetime.today().strftime("%Y-%m-%d")
    report_date = datetime.strptime(date_arg, "%Y-%m-%d")
    out_name    = f"Successful_IPS_Transaction_for_{report_date.strftime('%B_%d_%Y')}.xlsx"
    generate_report(input_file, f"/mnt/user-data/outputs/{out_name}", report_date)
