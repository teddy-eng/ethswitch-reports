import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
from datetime import datetime

NAME_MAP = {
    "Abay Bank":      "Abay",
    "Addis bank":     "Addis",
    "Ahadu Ebirr":    "Ahadu Ebirr",
    "Amhara Bank":    "Amhara",
    "Awash Bank":     "Awash ",
    "BOA":            "Abyssinia",
    "Berhan Bank":    "Berhan ",
    "Buna bank":      "Buna ",
    "CBE":            "CBE",
    "CBEBirr":        "CBE BIRR",
    "COOP":           "CBO",
    "Dashen Bank":    "Dashen ",
    "Dedebit":        "Dedebit",
    "Enat Bank":      "Enat ",
    "GOH BETOCH":     "GOH ",
    "Gadaa":          "Gadaa",
    "Global":         "Global",
    "HalalPay":       "HalalPay",
    "Hibret":         "Hibret",
    "Hijra Bank":     "Hijra ",
    "KAAFI":          "KAAFI ",
    "LIB":            "LIB",
    "MPESA":          "MPESA",
    "Nib Bank":       "NIB",
    "NibBirr":        "NibE Birr",
    "Nisir":          "Nisir",
    "Omo":            "Omo",
    "Oromia bank":    "Oromia ",
    "Rammis Bank":    "Rammis ",
    "Sidama Bank":    "Sidama ",
    "Siinqee Bank":   "Siinqee",
    "Siket":          "Siket",
    "Tsedey Bank":    "Tsedey ",
    "Tsehay Bank":    "Tsehay",
    "Vision Fund":    "Vision Fund",
    "Wegagen Bank":   "Wegagen",
    "Wegagen E-Birr": "Wegagen e-Birr",
    "ZamZam":         "ZamZam",
    "Zemen Bank":     "Zemen ",
    "telebirr":       "telebirr",
    "Coopay-E-Birr":  "Coopay-E-Birr",
    "H-Cash":         "H-Cash",
    "Kacha":          "Kacha",
    "Ahadu":          "Ahadu",
    "Dire MFI":       "Dire MFI",
    "Shabelle":       "Shabelle",
    "Vitabirr":       "Vitabirr",
    "Rays":           "Rays",
    "Saha":           "Saha",
}

PERMANENT_FIS = [
    "Abay", "Addis", "Ahadu", "Ahadu Ebirr", "Amhara", "Awash ",
    "Abyssinia", "Berhan ", "Buna ", "CBE", "CBE BIRR", "CBO",
    "Coopay-E-Birr", "Dashen ", "Dedebit", "Enat ", "GOH ", "Gadaa",
    "Global", "HalalPay", "H-Cash", "Hibret", "Hijra ", "KAAFI ",
    "Kacha", "LIB", "MPESA", "NIB", "NibE Birr", "Nisir", "Omo",
    "Oromia ", "Rammis ", "Rays", "Saha", "Shabelle", "Sidama ",
    "Siinqee", "Siket", "Tsedey ", "Tsehay", "Vision Fund",
    "Vitabirr", "Wegagen", "Wegagen e-Birr", "ZamZam", "Zemen ",
    "telebirr",
]

ERROR_COLS = ["ERRR", "MISS", "FRAD", "UNFN", "BLCK", "Other", "NullError"]

RED    = "FF0000"
YELLOW = "FFFF00"
GREEN  = "00B050"
BLACK  = "000000"
WHITE  = "FFFFFF"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def success_color(rate):
    if rate is None or rate >= 0.987:
        return make_fill(GREEN)
    elif rate > 0.96:
        return make_fill(YELLOW)
    else:
        return make_fill(RED)

def generate_p2p_report(error_path, success_path, output_path, report_date):
    err_df = pd.read_excel(error_path)
    err_df = err_df[err_df["Bank Name"].notna() & (err_df["Bank Name"] != "TOTAL")]
    err_df["DISPLAY"] = err_df["Bank Name"].map(NAME_MAP).fillna(err_df["Bank Name"])

    suc_df = pd.read_excel(success_path)
    suc_df["DISPLAY"] = suc_df["BANK_NAME"].map(NAME_MAP).fillna(suc_df["BANK_NAME"])

    def to_int(val):
        if pd.isna(val): return 0
        return int(str(val).replace(",", ""))

    # Build error data
    error_data = {}
    unmapped = []
    for _, r in err_df.iterrows():
        name = r["DISPLAY"]
        raw  = str(r["Bank Name"]).strip()
        if raw not in NAME_MAP:
            unmapped.append(raw)
        error_data[name] = {col: to_int(r.get(col, 0)) for col in ERROR_COLS}
        error_data[name]["Total Errors"] = to_int(r.get("Total Errors", 0))

    # Build success data (ISSUER_TXN_COUNT only)
    success_data = {}
    for _, r in suc_df.iterrows():
        name = r["DISPLAY"]
        cnt  = to_int(r.get("ISSUER_TXN_COUNT", 0))
        success_data[name] = success_data.get(name, 0) + cnt

    # Auto-promote all seen FIs to permanent list
    for name in list(error_data.keys()) + list(success_data.keys()):
        if name not in PERMANENT_FIS:
            PERMANENT_FIS.append(name)

    all_fis = sorted(set(list(error_data.keys()) + list(success_data.keys()) + PERMANENT_FIS),
                     key=lambda x: x.strip().lower())

    # Compute stats per FI
    fi_stats = {}
    for fi in all_fis:
        errs         = error_data.get(fi, {col: 0 for col in ERROR_COLS + ["Total Errors"]})
        unfn         = errs.get("UNFN", 0)
        issuer       = success_data.get(fi, 0)                         # raw ISSUER count
        total_succ   = issuer + unfn                                    # row14 = UNFN + Successful
        total_decline= errs.get("Total Errors", 0) - unfn              # row12 excl UNFN
        total_txn    = total_succ + total_decline                       # row15
        rate         = (total_succ / total_txn) if total_txn > 0 else None
        fi_stats[fi] = {
            "errors": errs, "issuer": issuer, "unfn": unfn,
            "total_succ": total_succ, "total_decline": total_decline,
            "total": total_txn, "rate": rate,
        }

    if unmapped:
        print("\n⚠️  UNMAPPED FIs:", unmapped)
    else:
        print("✅ All FIs matched.")

    date_str = report_date.strftime("%B %d,%Y")
    wb = Workbook()

    # ══════════════════════════════════════════════
    # SHEET 1
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Sheet1"

    n      = len(all_fis)
    last_c = get_column_letter(n + 2)   # +1 for label col A, +1 for Total col

    # Title rows 1-3
    ws1.merge_cells(f"A1:{last_c}3")
    ws1["A1"] = f"EthSwitch \n{date_str}  IPS-P2P Transaction Success Rate Report"
    ws1["A1"].font      = Font(bold=True, size=14)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in [1, 2, 3]:
        ws1.row_dimensions[r].height = 14.25

    # Row 4: header
    ws1.row_dimensions[4].height = 27.0
    ws1["A4"] = "Declined RC/Bank Name"
    ws1["A4"].font      = Font(bold=True, size=11)
    ws1["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1["A4"].border    = thin_border()
    ws1.column_dimensions["A"].width = 25.0

    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        cell = ws1[f"{col}4"]
        cell.value     = fi
        cell.font      = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()

    total_col = get_column_letter(n + 2)
    ws1[f"{total_col}4"].value     = "Total"
    ws1[f"{total_col}4"].font      = Font(bold=True, size=11)
    ws1[f"{total_col}4"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"{total_col}4"].border    = thin_border()

    # Rows 5-11: error codes
    for ri, label in enumerate(ERROR_COLS):
        r = 5 + ri
        ws1.row_dimensions[r].height = 15.0
        ws1[f"A{r}"] = label
        ws1[f"A{r}"].font   = Font(size=11)
        ws1[f"A{r}"].border = thin_border()
        for i, fi in enumerate(all_fis):
            col  = get_column_letter(i + 2)
            cell = ws1[f"{col}{r}"]
            cell.value  = fi_stats[fi]["errors"].get(label, 0)
            cell.font   = Font(size=12)
            cell.border = thin_border()
        # Total formula
        first = "B"
        last  = get_column_letter(n + 1)
        ws1[f"{total_col}{r}"].value        = f"=SUM({first}{r}:{last}{r})"
        ws1[f"{total_col}{r}"].number_format = "#,##0"
        ws1[f"{total_col}{r}"].border        = thin_border()

    # Row 12: Total Decline Txn (excludes UNFN = row 8)
    ws1.row_dimensions[12].height = 15.0
    ws1["A12"] = "Total Decline Txn"
    ws1["A12"].font = Font(bold=True, size=11); ws1["A12"].border = thin_border()
    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        cell = ws1[f"{col}12"]
        cell.value  = f"={col}5+{col}6+{col}7+{col}9+{col}10+{col}11"
        cell.font   = Font(size=12)
        cell.border = thin_border()
    ws1[f"{total_col}12"].value        = f"={total_col}5+{total_col}6+{total_col}7+{total_col}9+{total_col}10+{total_col}11"
    ws1[f"{total_col}12"].border        = thin_border()

    # Row 13: Successful Txn (ISSUER_TXN_COUNT — plain numbers)
    ws1.row_dimensions[13].height = 15.0
    ws1["A13"] = "Successful Txn"
    ws1["A13"].font = Font(bold=True, size=11); ws1["A13"].border = thin_border()
    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        cell = ws1[f"{col}13"]
        cell.value  = fi_stats[fi]["issuer"]
        cell.font   = Font(size=12)
        cell.border = thin_border()
    ws1[f"{total_col}13"].value         = f"=SUM(B13:{get_column_letter(n+1)}13)"
    ws1[f"{total_col}13"].number_format = "#,##0"
    ws1[f"{total_col}13"].border         = thin_border()

    # Row 14: Total Successful Txn = UNFN (row8) + Successful Txn (row13)
    ws1.row_dimensions[14].height = 15.0
    ws1["A14"] = "Total Succesful Txn"
    ws1["A14"].font = Font(bold=True, size=11); ws1["A14"].border = thin_border()
    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        cell = ws1[f"{col}14"]
        cell.value        = f"=SUM({col}8+{col}13)"
        cell.font         = Font(size=12)
        cell.number_format = "#,##0"
        cell.alignment    = Alignment(horizontal="center")
        cell.border       = thin_border()
    ws1[f"{total_col}14"].value         = f"=SUM({total_col}8+{total_col}13)"
    ws1[f"{total_col}14"].number_format = "#,##0"
    ws1[f"{total_col}14"].border         = thin_border()

    # Row 15: Total Txn = Total Decline (row12) + Total Successful (row14)
    num_fmt = '_(* #,##0_);_(* \\(#,##0\\);_(* \\-??_);_(@_)'
    ws1.row_dimensions[15].height = 15.0
    ws1["A15"] = "Total Txn"
    ws1["A15"].font = Font(bold=True, size=11); ws1["A15"].border = thin_border()
    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        cell = ws1[f"{col}15"]
        cell.value        = f"=SUM({col}12+{col}14)"
        cell.font         = Font(size=12)
        cell.number_format = num_fmt
        cell.border       = thin_border()
    ws1[f"{total_col}15"].value         = f"=SUM({total_col}12+{total_col}14)"
    ws1[f"{total_col}15"].number_format = num_fmt
    ws1[f"{total_col}15"].border         = thin_border()

    # Row 16: IPS-P2P SuccessRate = Total Successful (row14) / Total Txn (row15)
    ws1.row_dimensions[16].height = 15.0
    ws1["A16"] = "IPS-P2P SuccessRate"
    ws1["A16"].font = Font(bold=True, size=11); ws1["A16"].border = thin_border()
    for i, fi in enumerate(all_fis):
        col  = get_column_letter(i + 2)
        rate = fi_stats[fi]["rate"]
        cell = ws1[f"{col}16"]
        cell.value        = f"={col}14/{col}15" if fi_stats[fi]["total"] > 0 else 1
        cell.font         = Font(bold=True, size=11)
        cell.number_format = "0.00%"
        cell.alignment    = Alignment(horizontal="center")
        cell.fill         = success_color(rate)
        cell.border       = thin_border()

    total_succ_all = sum(fi_stats[fi]["total_succ"] for fi in all_fis)
    total_all      = sum(fi_stats[fi]["total"]      for fi in all_fis)
    total_rate     = total_succ_all / total_all if total_all > 0 else None
    tc16 = ws1[f"{total_col}16"]
    tc16.value        = f"={total_col}14/{total_col}15"
    tc16.font         = Font(bold=True, size=11)
    tc16.number_format = "0.00%"
    tc16.alignment    = Alignment(horizontal="center")
    tc16.fill         = success_color(total_rate)
    tc16.border       = thin_border()

    # Legend rows 18-20
    ws1["B18"] = "Range";      ws1["B18"].font = Font(bold=True, size=11)
    ws1["C18"] = ">=98.7%";    ws1["C18"].number_format = "0%"; ws1["C18"].font = Font(size=11)
    ws1["D18"] = "Green";      ws1["D18"].fill = make_fill(GREEN); ws1["D18"].font = Font(size=11)
    ws1["C19"] = "96%<X<98.7%"; ws1["C19"].font = Font(size=11)
    ws1["D19"] = "Yellow";     ws1["D19"].fill = make_fill(YELLOW); ws1["D19"].font = Font(size=11)
    ws1["C20"] = "<= 96%";     ws1["C20"].font = Font(size=11)
    ws1["D20"] = "Red";        ws1["D20"].fill = make_fill(RED); ws1["D20"].font = Font(size=11)

    # ══════════════════════════════════════════════
    # SHEET 2 — Ranked summary, highest → lowest
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("Sheet2")

    ws2.merge_cells("J1:K1")
    ws2.merge_cells("J2:K2")
    ws2["J2"] = date_str
    ws2["J2"].font      = Font(bold=True, size=14)
    ws2["J2"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[2].height = 17.25

    ws2["J3"] = "FI\u2019s Name"
    ws2["K3"] = "IPS P2P Success Rate"
    ws2["J3"].font = Font(bold=True, size=14)
    ws2["K3"].font = Font(bold=True, size=14)
    ws2.row_dimensions[3].height = 18.0
    ws2.column_dimensions["J"].width = 15.71
    ws2.column_dimensions["K"].width = 25.29

    # Sort: highest rate first (None → 100%)
    sorted_fis = sorted(all_fis, key=lambda fi: -(fi_stats[fi]["rate"] if fi_stats[fi]["rate"] is not None else 1.0))

    for i, fi in enumerate(sorted_fis):
        row  = 4 + i
        rate = fi_stats[fi]["rate"]
        disp = rate if rate is not None else 1.0

        ws2[f"J{row}"] = fi
        ws2[f"J{row}"].font      = Font(bold=True, size=11)
        ws2[f"J{row}"].alignment = Alignment(horizontal="center")
        ws2[f"J{row}"].border    = thin_border()

        cell = ws2[f"K{row}"]
        cell.value        = disp
        cell.font         = Font(bold=True, size=11)
        cell.number_format = "0.00%"
        cell.alignment    = Alignment(horizontal="center")
        cell.fill         = success_color(disp)
        cell.border       = thin_border()

    # Total row
    total_r = 4 + len(sorted_fis)
    ws2[f"J{total_r}"] = "Total"
    ws2[f"J{total_r}"].font      = Font(bold=True, size=11)
    ws2[f"J{total_r}"].alignment = Alignment(horizontal="center")
    ws2[f"J{total_r}"].border    = thin_border()

    total_disp = total_rate if total_rate is not None else 1.0
    tc2 = ws2[f"K{total_r}"]
    tc2.value        = total_disp
    tc2.font         = Font(bold=True, size=11)
    tc2.number_format = "0.00%"
    tc2.alignment    = Alignment(horizontal="center")
    tc2.fill         = success_color(total_disp)
    tc2.border       = thin_border()

    wb.save(output_path)
    print(f"✅ P2P Success Rate Report saved → {output_path}")

if __name__ == "__main__":
    error_path   = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/IPS_Error_Report_2026-07-08_to_2026-07-08.xlsx"
    success_path = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/uploads/IPS_success_for_source_and_destination_14_.xlsx"
    date_arg     = sys.argv[3] if len(sys.argv) > 3 else "2026-07-08"
    report_date  = datetime.strptime(date_arg, "%Y-%m-%d")
    out_name     = f"IPS-P2P_Success_Rate_Report_for_{report_date.strftime('%B_%d_%Y')}.xlsx"
    output_file  = f"/mnt/user-data/outputs/{out_name}"
    generate_p2p_report(error_path, success_path, output_file, report_date)
