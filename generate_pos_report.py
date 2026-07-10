import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
from datetime import datetime

# ── Name mapping: raw name → display name ──
ISSUER_NAME_MAP = {
    "Abay Bank":            "Abay",
    "Abyssinia Bank":       "Abyssinia",
    "Addis Int Bank":       "Addis",
    "Ahadu Bank":           "Ahadu",
    "Amhara Bank":          "Amhara",
    "Awash Bank":           "Awash",
    "Berhan Bank":          "Berhan",
    "Bunna Bank":           "Bunna",
    "Bunna Int Bank":       "Bunna",
    "CBO Switch":           "CBO",
    "Commercial Bank":      "CBE",
    "Coop Bank of Oromia":  "CBO",
    "Sinqee Bank":          "Siinqee",
    "Dashen Bank":          "Dashen",
    "Enat Bank":            "Enat",
    "Hibret Bank":          "Hibret",
    "Hijra Bank":           "Hijira",
    "Lion Int Bank":        "Lion",
    "Nib Int Bank":         "NIB",
    "Oromia Bank":          "Oromiya ",
    "Siket Bank":           "Siket Bank",
    "Tsedey Bank":          "Tsedey",
    "Tsehay Bank":          "Tsehay",
    "Wegagen Bank":         "Wegagen",
    "ZamZam Bank":          "ZamZam",
    "Zemen Bank":           "Zemen",
    "Gadaa Bank":           "Gadaa",
    "Global Bank":          "Global",
    "Rammis Bank":          "Rammis",
}

ACQUIRER_NAME_MAP = {
    "Abay Bank":        "Abay",
    "Abyssinia Bank":   "Abyssinia",
    "Addis Int Bank":   "Addis",
    "Amhara Bank":      "Amhara",
    "Arifpay":          "Arif pay",
    "Awash Bank":       "Awash",
    "Berhan Bank":      "Berhan",
    "CBO Switch":       "CBO",
    "Commercial Bank":  "CBE",
    "Dashen Bank":      "Dashen",
    "Hibret Bank":      "Hibret",
    "Nib Int Bank":     "NIB",
    "Oromia Bank":      "Oromiya ",
    "Santimpay":        "Santim Pay",
    "Siket Bank":       "Siket Bank",
    "Sinqee Bank":      "Siinqee",
    "Wegagen Bank":     "Wegagen",
    "YagoutPay":        "Yagot Pay",
    "Zemen Bank":       "Zemen",
    "Gadaa Bank":       "Gadaa",
    "Global Bank":      "Global",
    "Rammis Bank":      "Rammis",
    "Hijra Bank":       "Hijira",
    "Lion Int Bank":    "Lion",
    "Enat Bank":        "Enat",
    "Bunna Bank":       "Bunna",
    "Tsedey Bank":      "Tsedey",
    "Tsehay Bank":      "Tsehay",
    "ZamZam Bank":      "ZamZam",
    "Ahadu Bank":       "Ahadu",
}

# Permanent FIs always shown even with no data
PERMANENT_FIS = [
    "Abay", "Abyssinia", "Addis", "Ahadu", "Amhara", "Arif pay",
    "Awash", "Berhan", "Bunna", "CBE", "CBO", "Dashen", "Enat",
    "Gadaa", "Global", "Hibret", "Hijira", "Lion", "NIB", "Oromiya ",
    "Rammis", "Santim Pay", "Siinqee", "Siket Bank", "Tsedey",
    "Tsehay", "Wegagen", "Yagot Pay", "ZamZam", "Zemen",
]

GREEN = "A9D08E"
WHITE = "FFFFFF"
BLACK = "000000"
YELLOW = "FFFF00"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def generate_pos_report(issuer_path, acquirer_path, output_path, report_date):
    iss_df = pd.read_excel(issuer_path)
    acq_df = pd.read_excel(acquirer_path)

    # Build issuer data
    issuer_data = {}
    unmapped = []
    for _, r in iss_df.iterrows():
        raw = str(r["ISSUER_BANKS"]).strip()
        name = ISSUER_NAME_MAP.get(raw)
        if name is None:
            unmapped.append(f"ISSUER:{raw}")
            name = raw
        cnt = int(r["POS_PURCHASE"]) if pd.notna(r.get("POS_PURCHASE")) else 0
        amt = float(r["AMOUNT"]) if pd.notna(r.get("AMOUNT")) else 0.0
        if name in issuer_data:
            issuer_data[name]["cnt"] += cnt
            issuer_data[name]["amt"] += amt
        else:
            issuer_data[name] = {"cnt": cnt, "amt": amt}

    # Build acquirer data
    acquirer_data = {}
    for _, r in acq_df.iterrows():
        raw = str(r["ACQUIRER_BANKS"]).strip()
        name = ACQUIRER_NAME_MAP.get(raw)
        if name is None:
            unmapped.append(f"ACQUIRER:{raw}")
            name = raw
        cnt = int(r["POS_PURCHASE"]) if pd.notna(r.get("POS_PURCHASE")) else 0
        amt = float(r["Amount"]) if pd.notna(r.get("Amount")) else 0.0
        acquirer_data[name] = {"cnt": cnt, "amt": amt}

    # Auto-promote all seen FIs to permanent list
    for name in list(issuer_data.keys()) + list(acquirer_data.keys()):
        if name not in PERMANENT_FIS:
            PERMANENT_FIS.append(name)

    # Build master sorted FI list
    all_fis = sorted(set(list(issuer_data.keys()) + list(acquirer_data.keys()) + PERMANENT_FIS),
                     key=lambda x: x.strip().lower())

    if unmapped:
        print(f"\n⚠️  UNMAPPED FIs: {unmapped}")
    else:
        print("✅ All FIs matched.")

    date_str = report_date.strftime("%B %d,%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Header rows ──
    ws.merge_cells("C1:G1")
    ws["C1"] = "EthSwitch S.C."
    ws["C1"].font = Font(bold=True, size=14)
    ws["C1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25.15

    ws.merge_cells("C2:G2")
    ws["C2"] = report_date
    ws["C2"].number_format = "d-mmm-yy"
    ws["C2"].font = Font(bold=True, size=14)
    ws["C2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 24.0

    ws.merge_cells("C3:G3")
    ws["C3"] = f"Successful POS Interoperable Transactions of {date_str}"
    ws["C3"].font = Font(bold=True, size=14, color=BLACK)
    ws["C3"].fill = make_fill(WHITE)
    ws["C3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 33.0

    # Row 4: column headers
    ws.row_dimensions[4].height = 49.15
    ws.merge_cells("C4:C5"); ws["C4"] = "BANK"
    ws["C4"].font = Font(size=14, color=BLACK)
    ws["C4"].fill = make_fill(GREEN)
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C4"].border = thin_border()

    ws.merge_cells("D4:E4"); ws["D4"] = "Successful No. POS Purchase Transactions "
    ws["D4"].font = Font(size=14, color=BLACK)
    ws["D4"].fill = make_fill(GREEN)
    ws["D4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D4"].border = thin_border()

    ws.merge_cells("F4:G4"); ws["F4"] = "Total Value of Transactions"
    ws["F4"].font = Font(size=14, color=BLACK)
    ws["F4"].fill = make_fill(GREEN)
    ws["F4"].number_format = "#,##0.00"
    ws["F4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["F4"].border = thin_border()

    # Row 5: sub-headers
    ws.row_dimensions[5].height = 30.0
    for col, label in [("D", "As ISSUER"), ("E", "As ACQUIRER"), ("F", "As ISSUER"), ("G", "As ACQUIRER")]:
        ws[f"{col}5"] = label
        ws[f"{col}5"].font = Font(size=14, color=BLACK)
        ws[f"{col}5"].fill = make_fill(GREEN)
        ws[f"{col}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}5"].border = thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 8.86
    ws.column_dimensions["C"].width = 16.57
    ws.column_dimensions["D"].width = 19.71
    ws.column_dimensions["E"].width = 20.29
    ws.column_dimensions["F"].width = 19.71
    ws.column_dimensions["G"].width = 21.29

    # ── Data rows starting at row 6 ──
    data_start = 6
    for i, fi in enumerate(all_fis):
        r = data_start + i
        iss = issuer_data.get(fi, {"cnt": None, "amt": None})
        acq = acquirer_data.get(fi, {"cnt": None, "amt": None})

        ws[f"C{r}"] = fi
        ws[f"C{r}"].font = Font(size=14)
        ws[f"C{r}"].border = thin_border()

        # Issuer count
        cell_d = ws[f"D{r}"]
        cell_d.value = iss["cnt"] if iss["cnt"] is not None else None
        cell_d.font = Font(size=14)
        cell_d.alignment = Alignment(horizontal="center")
        cell_d.number_format = "#,##0"
        cell_d.border = thin_border()

        # Acquirer count
        cell_e = ws[f"E{r}"]
        cell_e.value = acq["cnt"] if acq["cnt"] is not None else None
        cell_e.font = Font(size=14)
        cell_e.alignment = Alignment(horizontal="center")
        cell_e.border = thin_border()

        # Issuer amount
        cell_f = ws[f"F{r}"]
        cell_f.value = iss["amt"] if iss["amt"] is not None else None
        cell_f.font = Font(size=14)
        cell_f.alignment = Alignment(horizontal="center")
        cell_f.number_format = "#,##0"
        cell_f.border = thin_border()

        # Acquirer amount
        cell_g = ws[f"G{r}"]
        cell_g.value = acq["amt"] if acq["amt"] is not None else None
        cell_g.font = Font(size=14)
        cell_g.alignment = Alignment(horizontal="center")
        cell_g.number_format = "#,##0"
        cell_g.border = thin_border()

    # ── TOTAL row ──
    total_row = data_start + len(all_fis)
    last_data = total_row - 1

    ws[f"C{total_row}"] = "TOTAL"
    ws[f"C{total_row}"].font = Font(bold=True, size=14, color=BLACK)
    ws[f"C{total_row}"].fill = make_fill(GREEN)
    ws[f"C{total_row}"].alignment = Alignment(horizontal="center")
    ws[f"C{total_row}"].border = thin_border()

    for col, fmt in [("D", "#,##0"), ("E", "#,##0"), ("F", "#,##0"), ("G", "#,##0")]:
        cell = ws[f"{col}{total_row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{last_data})"
        cell.font = Font(bold=True, size=14, color=BLACK)
        cell.fill = make_fill(GREEN)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = fmt
        cell.border = thin_border()

    wb.save(output_path)
    print(f"✅ POS Report saved → {output_path}")
    if unmapped:
        print("⚠️  Review yellow-highlighted rows before sending to management!")

if __name__ == "__main__":
    issuer_path  = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/Issuer.xlsx"
    acquirer_path= sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/uploads/Acquirer.xlsx"
    date_arg     = sys.argv[3] if len(sys.argv) > 3 else "2026-07-09"
    report_date  = datetime.strptime(date_arg, "%Y-%m-%d")
    out_name     = f"POS_Successful_With_Value_{report_date.strftime('%B_%d_%Y')}.xlsx"
    output_file  = f"/mnt/user-data/outputs/{out_name}"
    generate_pos_report(issuer_path, acquirer_path, output_file, report_date)
