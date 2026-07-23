"""
POS Success Rate Report generator for EthSwitch — app module.

Drop this file into the same repo as app.py, generate_ips_report.py, etc.
Exposes: generate_pos_success_rate_report(input_path, output_path, report_date)
to match the calling convention already used by the other report generators
in this app (e.g. generate_qr_report, generate_p2p_report).

Business logic (confirmed against real template/raw data pairs on 2026-07-17):
  - Only rows with TRANS_TYPE == "POS purchase" count (POS balance inquiry excluded).
  - RESP == -1  -> Successful Pos T
  - Cardholder-related decline codes (treated as "successful" for success-rate purposes):
        821, 901, 904, 906, 911, 914, 915
    For Wegagen Bank (WGB) ONLY, code 503 is additionally treated as cardholder-related.
  - success rate = (Successful Pos T + card holder rel dec) / total pos t
        total pos t = Successful Pos T + Total Decline (ALL decline codes)
  - Coloring on success rate:
        >=96%             -> green   (FF00B050)
        86% - 95.999%     -> yellow  (FFFFFF00)
        79% - 85.999%     -> light yellow / orange (FFFFC000)
        <79%              -> red     (FFFF0000)
  - New banks appearing in the raw ISSUER column that aren't yet in BANK_COLUMNS
    are added as new report columns automatically going forward.
  - New RESP codes not yet in RESPONSE_CODE_ORDER are added as extra rows
    automatically going forward.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG — edit here if a new bank joins the switch or business logic changes
# ---------------------------------------------------------------------------

# (template column label, [matching ISSUER names in the raw export])
BANK_COLUMNS = [
    ("Abay",   ["Abay Bank"]),
    ("BOA",    ["Abyssinia Bank"]),
    ("Adiss",  ["Addis Int Bank"]),
    ("Ahadu",  ["Ahadu Bank"]),
    ("Amhara", ["Amhara Bank"]),
    ("AWASH",  ["Awash Bank"]),
    ("BRB",    ["Berhan Bank"]),
    ("BIB",    ["Bunna Bank", "Bunna Int Bank"]),
    ("CBE",    ["Commercial Bank"]),
    ("CBO",    ["CBO Switch", "Coop Bank of Oromia"]),
    ("Dashen", ["Dashen Bank"]),
    ("Enat",   ["Enat Bank"]),
    ("Hibret", ["Hibret Bank"]),
    ("Lion",   ["Lion Int Bank"]),
    ("NIB",    ["Nib Int Bank"]),
    ("OB",     ["Oromia Bank"]),
    ("Sinqee", ["Sinqee Bank"]),
    ("Tsehay", ["Tsehay Bank"]),
    ("WGB",    ["Wegagen Bank"]),
    ("Zamzam", ["ZamZam Bank"]),
    ("Zemen",  ["Zemen Bank"]),
    # New banks added 2026-07-17 per Teddy's decision — append new ones below.
    ("Hijra",  ["Hijra Bank"]),
    ("Sidama", ["Sidama Bank"]),
    ("Siket",  ["Siket Bank"]),
    ("Tsedey", ["Tsedey Bank"]),
]

# Known decline codes (everything except the fixed tail), always displayed ascending.
KNOWN_CODES = [
    503, 801, 802, 803, 805, 806, 811, 820, 821, 827, 835, 882, 886,
    901, 904, 905, 906, 911, 912, 914, 917, 928, 939, 953, 959,
]
# These three always stay at the very end, in this exact order, no matter what.
RESPONSE_CODE_TAIL = [997, 987, 915]

CARDHOLDER_CODES = {821, 901, 904, 906, 911, 912, 914, 915}
WEGAGEN_LABEL = "WGB"
WEGAGEN_EXTRA_CODE = 503

# Static remark glossary (carried over verbatim from the template).
REMARKS = {
    503: "Not valid EMV transaction",
    504: "Card Operating rule does Not allow this transaction",
    801: "Bank Core banking system (CBS) or Issuer Switch did not respond",
    802: "Bank is disconnected from  Ethswitch",
    804: "Card is not permitted by the system for transaction",
    805: None,
    812: "The message was received in wrong format which can not be parsable by the system",
    821: "Wrong PIN is entered, wrong PIN is entered 3 and more times",
    827: "Generic Response from the Bank (Exactly not known)",
    857: "Requested amount was out of range allowed by the issuer.",
    858: "Error related with Keys",
    862: "Wrong PIN is entered 3 times & card is not captured by ATM",
    873: "Card is unknown by Ethswitch",
    878: "Account is locked in CBS",
    886: "Card is not in active status",
    901: "Customer enter invalid PIN or wrong PIN",
    902: "The transaction is cannot be processed by the system due to some format error",
    904: "Wrong PIN is entered 3 times & card is captured by ATM",
    905: "Card is not found in Data base",
    906: "Card Has Expired",
    909: "The card is not valid or cannot be used for transaction and captured by the ATM",
    911: "Maximum limit of amount a customer can withdraw is reached",
    912: "Maximum limit of amount a customer can withdraw is exceeded",
    913: "Transaction Type Not Supported By Institution",
    914: "wrong Account linked to card",
    915: "Customer wants to withdraw cash more than what he has in the account",
    917: "ATM or POS transaction amount limit exceeded",
    939: "Unknown response code responded by the Bank",
    952: "This Response code is sent when transaction is done using fallback method",
    959: "System malfunction",
    979: "Customer has savings account but select checking account while performing transaction or vice versa",
    988: "Service not available at the  time when transaction is performed",
}

GREEN, YELLOW, LIGHT_YELLOW, RED = "FF00B050", "FFFFFF00", "FFFFC000", "FFFF0000"
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _rate_color(rate_pct):
    if rate_pct >= 96:
        return GREEN
    elif rate_pct >= 86:
        return YELLOW
    elif rate_pct >= 79:
        return LIGHT_YELLOW
    else:
        return RED


def generate_pos_success_rate_report(input_path, output_path, report_date):
    """
    input_path: path to the raw SmartVista POS transaction export (.xlsx)
    output_path: where to save the generated report (.xlsx)
    report_date: a datetime object for the report's date
    """
    df = pd.read_excel(input_path)
    df = df[df["TRANS_TYPE"] == "POS purchase"].copy()

    known_codes = set(KNOWN_CODES) | set(RESPONSE_CODE_TAIL) | {-1}
    unknown_codes = sorted(set(df["RESP"].dropna().unique()) - known_codes)
    ascending_part = sorted(set(KNOWN_CODES) | set(unknown_codes))
    code_order = ascending_part + RESPONSE_CODE_TAIL

    mapped_issuers = {i for _, issuers in BANK_COLUMNS for i in issuers}
    raw_issuers = set(df["ISSUER"].dropna().unique())
    unmapped = raw_issuers - mapped_issuers
    if unmapped:
        # Auto-add unmapped issuers as their own new column so nothing is silently dropped.
        for issuer in sorted(unmapped):
            BANK_COLUMNS.append((issuer.replace(" Bank", "").replace(" Int", ""), [issuer]))

    counts = {}
    for label, issuers in BANK_COLUMNS:
        sub = df[df["ISSUER"].isin(issuers)]
        counts[label] = sub["RESP"].value_counts()

    n_banks = len(BANK_COLUMNS)
    first_data_row = 6
    last_data_row = first_data_row + len(code_order) - 1
    total_decline_row = last_data_row + 1
    successful_row = total_decline_row + 1
    success_rate_row = successful_row + 1
    cardholder_row = success_rate_row + 1
    total_succ_row = cardholder_row + 1
    total_post_row = total_succ_row + 1

    label_col = 1
    first_bank_col = 2
    last_bank_col = first_bank_col + n_banks - 1
    total_col = last_bank_col + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    title_str = f"{report_date.strftime('%B')} {report_date.day},{report_date.year} Pos Transaction Decline Response Summary"

    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=total_col)
    tcell = ws.cell(row=2, column=1, value=title_str)
    tcell.font = Font(name="Calibri", bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", wrap_text=True)

    header_font = Font(name="Calibri", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    hcell = ws.cell(row=5, column=label_col, value="RC/BANK NAME")
    hcell.font, hcell.alignment, hcell.border = header_font, header_align, BORDER
    for i, (label, _) in enumerate(BANK_COLUMNS):
        c = ws.cell(row=5, column=first_bank_col + i, value=label)
        c.font, c.alignment, c.border = header_font, header_align, BORDER
    c = ws.cell(row=5, column=total_col, value="Total")
    c.font, c.alignment, c.border = header_font, header_align, BORDER

    data_font = Font(name="Calibri", bold=True, size=11)
    for r_i, code in enumerate(code_order):
        row = first_data_row + r_i
        lc = ws.cell(row=row, column=label_col, value=code)
        lc.font, lc.border = data_font, BORDER
        for b_i, (label, _) in enumerate(BANK_COLUMNS):
            col = first_bank_col + b_i
            val = int(counts[label].get(code, 0))
            cell = ws.cell(row=row, column=col, value=val if val else None)
            cell.font, cell.border = data_font, BORDER
        first_letter, last_letter = get_column_letter(first_bank_col), get_column_letter(last_bank_col)
        tc = ws.cell(row=row, column=total_col, value=f"=SUM({first_letter}{row}:{last_letter}{row})")
        tc.font, tc.border = data_font, BORDER

    lc = ws.cell(row=total_decline_row, column=label_col, value="Total Decline")
    lc.font, lc.border = data_font, BORDER
    for b_i in range(n_banks):
        col = first_bank_col + b_i
        letter = get_column_letter(col)
        c = ws.cell(row=total_decline_row, column=col,
                    value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
        c.font, c.border = data_font, BORDER
    tc = ws.cell(row=total_decline_row, column=total_col,
                 value=f"=SUM({get_column_letter(first_bank_col)}{total_decline_row}:{get_column_letter(last_bank_col)}{total_decline_row})")
    tc.font, tc.border = data_font, BORDER

    lc = ws.cell(row=successful_row, column=label_col, value="Successful Pos T")
    lc.font, lc.border = data_font, BORDER
    for b_i, (label, _) in enumerate(BANK_COLUMNS):
        col = first_bank_col + b_i
        val = int(counts[label].get(-1, 0))
        c = ws.cell(row=successful_row, column=col, value=val)
        c.font, c.border = data_font, BORDER
    tc = ws.cell(row=successful_row, column=total_col,
                 value=f"=SUM({get_column_letter(first_bank_col)}{successful_row}:{get_column_letter(last_bank_col)}{successful_row})")
    tc.font, tc.border = data_font, BORDER

    code_row = {code: first_data_row + i for i, code in enumerate(code_order)}
    lc = ws.cell(row=cardholder_row, column=label_col, value="card holder rel dec")
    lc.font, lc.border = data_font, BORDER
    for b_i, (label, _) in enumerate(BANK_COLUMNS):
        col = first_bank_col + b_i
        letter = get_column_letter(col)
        codes_for_bank = set(CARDHOLDER_CODES)
        if label == WEGAGEN_LABEL:
            codes_for_bank.add(WEGAGEN_EXTRA_CODE)
        refs = "+".join(f"{letter}{code_row[c]}" for c in sorted(codes_for_bank) if c in code_row)
        c = ws.cell(row=cardholder_row, column=col, value=f"={refs}")
        c.font, c.border = data_font, BORDER
    tc = ws.cell(row=cardholder_row, column=total_col,
                 value=f"=SUM({get_column_letter(first_bank_col)}{cardholder_row}:{get_column_letter(last_bank_col)}{cardholder_row})")
    tc.font, tc.border = data_font, BORDER

    lc = ws.cell(row=total_succ_row, column=label_col, value="total succ")
    lc.font, lc.border = data_font, BORDER
    for b_i in range(n_banks):
        col = first_bank_col + b_i
        letter = get_column_letter(col)
        c = ws.cell(row=total_succ_row, column=col, value=f"={letter}{successful_row}+{letter}{cardholder_row}")
        c.font, c.border = data_font, BORDER
    tc = ws.cell(row=total_succ_row, column=total_col,
                 value=f"={get_column_letter(total_col)}{successful_row}+{get_column_letter(total_col)}{cardholder_row}")
    tc.font, tc.border = data_font, BORDER

    lc = ws.cell(row=total_post_row, column=label_col, value="total pos t")
    lc.font, lc.border = data_font, BORDER
    for b_i in range(n_banks):
        col = first_bank_col + b_i
        letter = get_column_letter(col)
        c = ws.cell(row=total_post_row, column=col, value=f"={letter}{total_decline_row}+{letter}{successful_row}")
        c.font, c.border = data_font, BORDER
    tc = ws.cell(row=total_post_row, column=total_col,
                 value=f"={get_column_letter(total_col)}{total_decline_row}+{get_column_letter(total_col)}{successful_row}")
    tc.font, tc.border = data_font, BORDER

    lc = ws.cell(row=success_rate_row, column=label_col, value="success rate")
    lc.font, lc.border = data_font, BORDER
    for b_i, (label, _) in enumerate(BANK_COLUMNS):
        col = first_bank_col + b_i
        letter = get_column_letter(col)
        extra = {WEGAGEN_EXTRA_CODE} if label == WEGAGEN_LABEL else set()
        succ = counts[label].get(-1, 0) + sum(counts[label].get(c, 0) for c in (CARDHOLDER_CODES | extra))
        total_decline_val = sum(counts[label].get(c, 0) for c in code_order)
        total_pos_val = counts[label].get(-1, 0) + total_decline_val
        rate_pct = (succ / total_pos_val * 100) if total_pos_val else 0
        c = ws.cell(row=success_rate_row, column=col, value=f"={letter}{total_succ_row}/{letter}{total_post_row}")
        c.number_format = "0%"
        c.font, c.border = data_font, BORDER
        c.fill = PatternFill(start_color=_rate_color(rate_pct), end_color=_rate_color(rate_pct), fill_type="solid")

    total_succ_all = sum(
        counts[label].get(-1, 0) + sum(
            counts[label].get(c, 0) for c in (CARDHOLDER_CODES | ({WEGAGEN_EXTRA_CODE} if label == WEGAGEN_LABEL else set()))
        ) for label, _ in BANK_COLUMNS
    )
    total_pos_all = sum(
        counts[label].get(-1, 0) + sum(counts[label].get(c, 0) for c in code_order) for label, _ in BANK_COLUMNS
    )
    total_rate_pct = (total_succ_all / total_pos_all * 100) if total_pos_all else 0
    tletter = get_column_letter(total_col)
    tc = ws.cell(row=success_rate_row, column=total_col, value=f"={tletter}{total_succ_row}/{tletter}{total_post_row}")
    tc.number_format = "0%"
    tc.font, tc.border = data_font, BORDER
    tc.fill = PatternFill(start_color=_rate_color(total_rate_pct), end_color=_rate_color(total_rate_pct), fill_type="solid")

    remark_header_row = total_post_row + 3
    remark_start_row = remark_header_row + 1
    hc1 = ws.cell(row=remark_header_row, column=1, value="Response Code")
    hc2 = ws.cell(row=remark_header_row, column=2, value="REMARK")
    for hc in (hc1, hc2):
        hc.font = Font(name="Arial", bold=True, size=11, color="FF000000")
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = BORDER
    for i, (code, remark) in enumerate(REMARKS.items()):
        row = remark_start_row + i
        rc = ws.cell(row=row, column=1, value=code)
        rc.font, rc.border = Font(name="Calibri", bold=True, size=11), BORDER
        rm = ws.cell(row=row, column=2, value=remark)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_col)
        rm.font = Font(name="Calibri", size=11, color="FF000000")
        rm.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(2, total_col + 1):
            ws.cell(row=row, column=col).border = BORDER
        rm.border = BORDER

    ws.column_dimensions["A"].width = 20
    for b_i in range(n_banks):
        ws.column_dimensions[get_column_letter(first_bank_col + b_i)].width = 9
    ws.column_dimensions[get_column_letter(total_col)].width = 9

    wb.save(output_path)
    return unknown_codes, sorted(unmapped)
