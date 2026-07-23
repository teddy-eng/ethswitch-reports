import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
from datetime import datetime

NAME_MAP = {
    "Awash Bank":     "Awash",
    "Dashen Bank":    "Dashen",
    "COOP":           "CBO",
    "CBE":            "CBE",
    "KAAFI":          "KAAFI",
    "Nib Bank":       "NIB",
    "Addis bank":     "Addis",
    "Wegagen Bank":   "Wegagen",
    "Abay Bank":      "Abay",
    "BOA":            "Abyssinia",
    "Buna bank":      "Bunna",
    "Siinqee Bank":   "Siinqee",
    "Wegagen E-Birr": "Wegagen E-Birr",
    "CBEBirr":        "CBE Birr",
    "MPESA":          "M-PESA",
    "Enat Bank":      "Enat Bank",
    "Siket":          "Siket",
    "Global":         "Global",
    "Vision Fund":    "Vision Fund",
    "Berhan Bank":    "Berhan Bank",
    "Ahadu":          "Ahadu",
    "Amhara Bank":    "Amhara",
    "Ahadu Ebirr":    "Ahadu Ebirr",
    "Coopay-E-Birr":  "Coopay-E-Birr",
    "GOH BETOCH":     "GOH BETOCH",
    "Hibret":         "Hibret",
    "Kacha":          "Kacha",
    "LIB":            "LIB",
    "Rammis Bank":    "Rammis Bank",
    "Saha":           "Saha",
    "telebirr":       "telebirr",
    "Tsedey Bank":    "Tsedey Bank",
    "Tsehay Bank":    "Tsehay",
    "Zemen Bank":     "Zemen",
    "Oromia bank":    "Oromia bank",
    "Gadaa":          "Gadaa",
    "Sidama Bank":    "Sidama Bank",
    "Nisir":          "Nisir",
    "Shabelle":       "Shabelle",
    "Omo":            "Omo Bank",
    "Dedebit":        "Dedebit MFI",
    "H-Cash":         "H-Cash",
    "Vitabirr":       "Vitabirr",
    "Hijra Bank":     "Hijra Bank",
    "Dire MFI":       "Dire MFI",
}

PERMANENT_FIS = [
    "Ahadu Ebirr", "Amhara", "Coopay-E-Birr", "GOH BETOCH",
    "Hibret", "Kacha", "LIB", "Rammis Bank", "Saha",
    "telebirr", "Tsedey Bank", "Tsehay", "Shabelle",
]

GREEN = "A9D08E"
WHITE = "FFFFFF"
BLACK = "000000"

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def generate_qr_report(input_path, output_path, report_date):
    df = pd.read_excel(input_path)

    unmapped = []
    rows = []
    seen_names = set()

    for _, r in df.iterrows():
        raw_name = str(r["BANK_NAME"]).strip()
        display_name = NAME_MAP.get(raw_name)
        if display_name is None:
            unmapped.append(raw_name)
            display_name = raw_name

        iss_cnt = int(r["ISSUER_TXN_COUNT"])       if pd.notna(r["ISSUER_TXN_COUNT"])     else 0
        iss_amt = float(r["ISSUER_TOTAL_AMOUNT"])   if pd.notna(r["ISSUER_TOTAL_AMOUNT"])  else 0.0
        acq_cnt = int(r["ACQUIRER_TXN_COUNT"])      if pd.notna(r["ACQUIRER_TXN_COUNT"])   else 0
        acq_amt = float(r["ACQUIRER_TOTAL_AMOUNT"]) if pd.notna(r["ACQUIRER_TOTAL_AMOUNT"]) else 0.0

        rows.append({
            "name": display_name,
            "iss_cnt": iss_cnt, "iss_amt": iss_amt,
            "acq_cnt": acq_cnt, "acq_amt": acq_amt,
        })
        seen_names.add(display_name)

        if display_name not in PERMANENT_FIS:
            PERMANENT_FIS.append(display_name)

    for fi in PERMANENT_FIS:
        if fi not in seen_names:
            rows.append({"name": fi, "iss_cnt": 0, "iss_amt": 0.0, "acq_cnt": 0, "acq_amt": 0.0})

    rows.sort(key=lambda x: x["name"].strip().lower())

    if unmapped:
        print(f"\n⚠️  NEW FI(s) DETECTED: {unmapped} — added to report automatically.\n")
    else:
        print("✅ All FIs matched.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    date_str = report_date.strftime("%B %d,%Y")

    ws.merge_cells("E2:I2"); ws["E2"] = "EthSwitch S.C."
    ws["E2"].font = Font(bold=True, size=12); ws["E2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("E3:I3"); ws["E3"] = "QR Report"
    ws["E3"].font = Font(bold=True, size=12); ws["E3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("E4:I4"); ws["E4"] = report_date
    ws["E4"].number_format = "MM/DD/YYYY"
    ws["E4"].font = Font(bold=True, size=12); ws["E4"].alignment = Alignment(horizontal="center")

    ws.merge_cells("E5:I5")
    ws["E5"] = f"Successful QR Interoperable Transactions for {date_str}"
    ws["E5"].font = Font(bold=True, size=11, color=BLACK)
    ws["E5"].fill = make_fill(WHITE); ws["E5"].alignment = Alignment(horizontal="center")

    ws.merge_cells("E6:E7"); ws["E6"] = "Bank"
    ws.merge_cells("F6:G6"); ws["F6"] = "As a Destination"
    ws.merge_cells("H6:I6"); ws["H6"] = "As a Source"
    ws["F7"] = "No.Transactions"; ws["G7"] = "Values"
    ws["H7"] = "No.Transactions"; ws["I7"] = "Values"

    for row in [6, 7]:
        for col in ["E", "F", "G", "H", "I"]:
            cell = ws[f"{col}{row}"]
            cell.fill = make_fill(GREEN)
            cell.font = Font(size=11, color=BLACK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

    data_start = 8
    for i, row in enumerate(rows):
        r = data_start + i
        ws[f"E{r}"] = row["name"]
        ws[f"E{r}"].fill = make_fill(GREEN)
        ws[f"E{r}"].font = Font(size=11, color=BLACK)
        ws[f"E{r}"].border = thin_border()
        ws[f"E{r}"].alignment = Alignment(horizontal="center")

        for col, val, fmt in [
            ("F", row["iss_cnt"], '#,##0'),
            ("G", row["iss_amt"], '#,##0.00'),
            ("H", row["acq_cnt"], '#,##0'),
            ("I", row["acq_amt"], '#,##0.00'),
        ]:
            cell = ws[f"{col}{r}"]
            cell.value = val
            cell.font = Font(size=11)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = fmt

    total_row = data_start + len(rows)
    last_data = total_row - 1
    ws[f"E{total_row}"] = "Total"
    ws[f"E{total_row}"].fill = make_fill(GREEN)
    ws[f"E{total_row}"].font = Font(size=11, color=BLACK)
    ws[f"E{total_row}"].border = thin_border()
    ws[f"E{total_row}"].alignment = Alignment(horizontal="center")

    for col, fmt in [("F", '#,##0'), ("G", '#,##0.00'), ("H", '#,##0'), ("I", '#,##0.00')]:
        cell = ws[f"{col}{total_row}"]
        cell.value = f"=SUM({col}{data_start}:{col}{last_data})"
        cell.font = Font(bold=True, size=11)
        cell.fill = make_fill(GREEN)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = fmt

    ws.column_dimensions["E"].width = 19.43
    ws.column_dimensions["F"].width = 18.14
    ws.column_dimensions["G"].width = 17.29
    ws.column_dimensions["H"].width = 18.71
    ws.column_dimensions["I"].width = 20.0

    wb.save(output_path)
    print(f"✅ QR Report saved → {output_path}")

if __name__ == "__main__":
    input_file  = sys.argv[1]
    date_arg    = sys.argv[2] if len(sys.argv) > 2 else datetime.today().strftime("%Y-%m-%d")
    report_date = datetime.strptime(date_arg, "%Y-%m-%d")
    out_name    = f"Successful_QR_Transaction_for_{report_date.strftime('%B_%d_%Y')}.xlsx"
    generate_qr_report(input_file, f"/mnt/user-data/outputs/{out_name}", report_date)
